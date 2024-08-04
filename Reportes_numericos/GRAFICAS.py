# graficas_personajes.py

from openpyxl import load_workbook
from collections import Counter
import matplotlib.pyplot as plt
import numpy as np
import os

def Altura_graf(frequency):
    plt.bar(frequency.keys(), frequency.values())
    plt.xlabel("Altura (cm)")
    plt.ylabel("Frequencia")
    plt.title("Alturas")
    plt.yticks(np.arange(0, max(frequency.values())+1, step=1))
    plt.grid(True)
    plt.show()

def Masa_graf(masas):
    plt.hist(masas, bins=10)
    plt.xlabel("Masa (lbs)")
    plt.ylabel("Frequencia")
    plt.title("Distribucion de masas")
    plt.grid(True)
    plt.show()

def genero_graf(etiquetas, cantidad):
    plt.pie(cantidad, labels=etiquetas, shadow=True)
    plt.axis("equal")
    plt.title("Distribucion de genero")
    plt.show()

def iris_graf(etiquetas, cantidad):
    plt.pie(cantidad, labels=etiquetas, shadow=True)
    plt.axis("equal")
    plt.title("Distribucion del color de ojo")
    plt.show()

def piel_graf(etiquetas, cantidad):
    plt.pie(cantidad, labels=etiquetas, shadow=True)
    plt.axis("equal")
    plt.title("Distribucion del color de piel")
    plt.show()

def generar_graficas():
    direccion_script = os.path.dirname(os.path.abspath(__file__))
    nombre_Excel = "personajes.xlsx"
    camino_excel = os.path.join(direccion_script, nombre_Excel)
    wb = load_workbook(camino_excel)
    ws = wb.active

    while True:
        try:
            opc = int(input("Bienvenido a la seccion de graficas, elija el numero correspondiente a la funcion para ejecutarla:\n***PERSONAJES***\n1.Distribucion de altura(Grafica de Barras)\n2.Distribucion de masas(Histograma)\n3.Distribucion de genero(Grafica de pastel)\n4.Distribucion de color de ojo(Grafica de pastel)\n5.Distribucion de color de piel(Grafica de pastel)\n6.Salir\n\n:"))
            if 1 <= opc <= 6:
                if opc == 1:
                    alturas = [cell.value for cell in ws["B"][1:]]
                    frequencia_altura = Counter(alturas)
                    Altura_graf(frequencia_altura) 
                elif opc == 2:
                    masas = [cell.value for cell in ws["C"][1:]]
                    Masa_graf(masas)
                elif opc == 3:
                    generos = [cell.value for cell in ws["H"][1:]]
                    cantidad_genero = Counter(generos)
                    etiquetas = cantidad_genero.keys()
                    cantidad = cantidad_genero.values()
                    genero_graf(etiquetas, cantidad)  
                elif opc == 4:
                    iris = [cell.value for cell in ws["F"][1:]]
                    cantidad_iris = Counter(iris)
                    etiquetas = cantidad_iris.keys()
                    cantidad = cantidad_iris.values()
                    iris_graf(etiquetas, cantidad) 
                elif opc == 5:
                    piel = [cell.value for cell in ws["E"][1:]]
                    cantidad_piel = Counter(piel)
                    etiquetas = cantidad_piel.keys()
                    cantidad = cantidad_piel.values()
                    piel_graf(etiquetas, cantidad) 
                else:
                    print('Bye bye...\n\n')
                    break
            else:
                print("Numero invalido, favor de poner numero entero del 1 al 6")
        except ValueError:
            print("Caracter invalido, favor de poner numero entero del 1 al 6")

if __name__ == "__main__":
    generar_graficas()
